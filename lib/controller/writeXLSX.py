from openpyxl import Workbook
from tkinter import *
from tkinter import filedialog

def writeXLSX(typeXLSX):    
    sourcePath = filedialog.askdirectory()

    wb = Workbook()
    ws = wb.active

    if typeXLSX == 'Operator':
        ws['A1'] = 'NAME'
        ws['B1'] = 'TELEFONE'
        ws['C1'] = 'CPF'
        ws['D1'] = 'EMAIL'
        ws['E1'] = 'LOGIN'
        ws['F1'] = 'PASSWORD'
        ws['G1'] = 'INACTIVE'
        fileName = '{}/ImportOperator.xlsx'.format(sourcePath)
    if typeXLSX == 'Product':
        ws['A1'] = 'PRODUCT'
        ws['B1'] = 'DESCRIPTION'
        ws['C1'] = 'INACTIVE'
        fileName = '{}/ImportProduct.xlsx'.format(sourcePath)
    if typeXLSX == 'Client':
        ws['A1'] = 'NAME'
        ws['B1'] = 'CPF'
        ws['C1'] = 'RG'
        ws['D1'] = 'BIRTH'
        ws['E1'] = 'SEX'
        ws['F1'] = 'EMAIL'
        ws['G1'] = 'CEP'
        ws['H1'] = 'ADDRESS'
        ws['I1'] = 'NUMBER'
        ws['J1'] = 'DISTRICT'
        ws['K1'] = 'CITY'
        ws['L1'] = 'STATE'
        ws['M1'] = 'TELEFONE'
        ws['N1'] = 'CELL'
        fileName = '{}/ImportClient.xlsx'.format(sourcePath)

    wb.save(fileName)
