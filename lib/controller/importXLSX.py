import openpyxl
import hashlib
from datetime import datetime
from controller.database.conf import *
from controller import function

def importXLSX(table, file_path):
    try:
        wb = openpyxl.load_workbook(filename=file_path)

        sheet = wb.active

        if table == 'Operator':
            return importOperator(sheet)
        if table == 'Product':
            return importProduct(sheet)
        if table == 'Client':
            return importClient(sheet)
    except:
        return 'Fail to read the file XLSX'


def importOperator(sheet):
    try:
        row = 2
        msg = ''

        conf = configurationElephant()
        cur = conf.cursor()

        # Pass for all columns
        for y in range(row, 1000):
            error = 0

            # 1
            temp = sheet.cell(row=y, column= 1).value
            if temp == None:
                break
            name = function.capitalizeWord(temp)
            
            # 2
            temp    = sheet.cell(row=y, column = 2).value
            if temp == None:
                if temp:
                    telefone = temp
            else:
                temp = function.validadeTelefone(str(temp))
                if temp:
                    telefone = sheet.cell(row=y, column = 2).value
                else:
                    msg += 'Erro row {}: The phone[{}] field was not filled in correctly.\n'.format(str(y), temp)
                    error += 1

            # 3
            temp         = sheet.cell(row=y, column = 3).value
            if temp == None:
                msg += 'Error row {}: The CPF[{}] field was not filled in correctly.\n'.format(str(y), temp)
                error += 1
            else:
                cpf     = function.validadeCPF(str(temp))
                if cpf:
                    cpf = temp
                    sql = "SELECT CPF FROM OPERATOR WHERE CPF = '{}'".format(cpf)
                    cur.execute(sql)
                    rows = cur.fetchall()
                    if rows:
                        msg += 'Erro linha {}: O CPF[{}] already exists in the database and cannot be inserted.\n'.format(str(y), cpf)
                        error += 1
                else:
                    msg += 'Erro row {}: The cpf[{}] field was not filled in correctly.\n'.format(str(y), cpf)
                    error += 1

            # 4
            if error == 0:
                temp       = sheet.cell(row=y, column = 4).value
                if temp == None:
                    msg += 'Erro linha {}: O campo email[{}] não está preenchido.\n'.format(str(y), temp)
                    error += 1
                else:
                    if function.validadeEmail(str(temp)):
                        email = function.capitalizeWord(temp)
                        sql = "SELECT EMAIL FROM OPERATOR WHERE EMAIL = '{}'".format(email)
                        cur.execute(sql)
                        rows = cur.fetchall()
                        if rows:
                            msg += 'Erro linha {}: O EMAIL[{}] already exists in the database and cannot be inserted.\n'.format(str(y), email)
                            error += 1
                    else:
                        msg += 'Erro row {}: The email[{}] field was not filled in correctly.\n'.format(str(y), email)
                        error += 1

            # 5
            if error == 0:
                temp       = sheet.cell(row=y, column = 5).value
                if temp == None:
                    msg += 'Error row {}: The login[{}] field was not filled in correctly.\n'.format(str(y), temp)
                    error += 1
                else:
                    login = function.capitalizeWord(str(temp))
                    sql = "SELECT LOGIN FROM OPERATOR WHERE LOGIN = '{}'".format(login)
                    cur.execute(sql)
                    rows = cur.fetchall()
                    if rows:
                        msg += 'Error row {}: The login[{}] already exists in the database and cannot be inserted.\n'.format(str(y), login)
                        error += 1

            # 6
            if error == 0:
                temp    = sheet.cell(row=y, column = 6).value
                if temp == None:
                    temp = 'Ch@nge123!'
                if function.validadePassword(temp):
                    password = temp
                    pass_hash = hashlib.sha1(password.encode('utf-8')).hexdigest()
                else:
                    msg += 'Erro row {}: The password[{}] field was not filled in correctly.\n'.format(str(y), email)
                    error += 1


            # 7
            if error == 0:
                temp    = sheet.cell(row=y, column = 7).value
                if temp == None:
                    inactive = 0
                if function.validadeNumber(str(temp)):
                    inactive = temp
                else:
                    incative = 0


            if error == 0:
                sql = "INSERT INTO OPERATOR(NAME, TELEFONE, CPF, EMAIL, LOGIN, PASSWORD, PASS_HASH, INACTIVE) VALUES('{}', {}, {}, '{}', '{}', '{}', '{}', {})".format(name, telefone, cpf, email, login, password, pass_hash, inactive)
                #debug
                # print(sql) 
                try:
                    cur.execute(sql)
                    conf.commit()
                    msg += 'Ok linha {}: O Operador[{}] foi cadastrado com sucesso.\n'.format(str(y), name)
                except:
                    msg += 'Erro linha {}: O Operador[{}] não foi possível ser cadastrado.\n'.format(str(y), name)
        fileName = "ImportOperator_{}.txt".format(str(datetime.now()))
        file = open(fileName, "w") 
        file.write(str(msg))
        file.close
        conf.close()
        return msg
    except Exception:
        msg = 'Fail to read de file XSLX'
        return msg
    
def importProduct(sheet):
    try:
        row = 2
        msg = ''

        conf = configurationElephant()
        cur = conf.cursor()

        # Pass for all columns
        for y in range(row, 1000):
            error = 0

            # 1
            tempProduct = sheet.cell(row=y, column= 1).value
            if tempProduct == None:
                break
            product = function.capitalizeWord(tempProduct)
            
            # 2
            tempDescription    = sheet.cell(row=y, column = 2).value
            if tempDescription == None:
                description = ''
            else:
                description = tempDescription

            

            # 3
            inactive    = sheet.cell(row=y, column = 3).value
            if inactive == None:
                inactive = 0

            if error == 0:
                sql = "INSERT INTO PRODUCT(PRODUCT, DESCRIPTION, INACTIVE) VALUES('{}', '{}', {})".format(product, description, inactive)
                #debug
                # print(sql) 
                try:
                    cur.execute(sql)
                    conf.commit()
                    msg += 'Ok row {}: The product [{}] was registered sucefull.\n'.format(str(y), product)
                except:
                    msg += 'Error row {}: The product[{}] was not registered with successful.\n'.format(str(y), product)
        fileName = "ImportProduct{}.txt".format(str(datetime.now()))
        file = open(fileName, "w") 
        file.write(str(msg))
        file.close
        conf.close()
        return msg
    except Exception:
        msg = 'Fail to read de file XSLX'
        return msg

def importClient(sheet):
    try:
        row = 2
        msg = ''

        conf = configurationElephant()
        cur = conf.cursor()

        # Pass for all columns
        for y in range(row, 1000):
            error = 0

            # 1
            name        = sheet.cell(row=y, column= 1).value
            if name == None:
                break

            # 2
            if error == 0:
                cpf       = sheet.cell(row=y, column = 2).value
                if cpf == None:
                    msg += 'Error row {}: The cell cpf[{}] wasnt preench.\n'.format(str(y), cpf)
                    error += 1
                else:
                    sql = "SELECT CPF FROM CLIENT WHERE CPF = '{}'".format(cpf)
                    cur.execute(sql)
                    rows = cur.fetchall()
                    if rows:
                        msg += 'Error row {}: The cpf[{}] was exists in the database and cannot register.\n'.format(str(y), cpf)
                        error += 1

            # 3
            if error == 0:
                rg       = sheet.cell(row=y, column = 3).value
                if rg == None:
                    msg += 'Error row {}: The cell rg[{}] wasnt preench.\n'.format(str(y), rg)
                    error += 1
                else:
                    sql = "SELECT RG FROM CLIENT WHERE RG = '{}'".format(rg)
                    cur.execute(sql)
                    rows = cur.fetchall()
                    if rows:
                        msg += 'Error row {}: The rg[{}] was exists in the database and cannot register.\n'.format(str(y), rg)
                        error += 1
                
            # 4
            birth       = sheet.cell(row=y, column= 4).value
            
            # 5
            sex         = sheet.cell(row=y, column= 5).value
            

            # 6
            if error == 0:
                email   = sheet.cell(row=y, column = 6).value
                if email == None:
                    msg += 'Error row {}: The cell email[{}] wasnt preench.\n'.format(str(y), email)
                    error += 1
                else:
                    sql = "SELECT EMAIL FROM CLIENT WHERE EMAIL = '{}'".format(email)
                    cur.execute(sql)
                    rows = cur.fetchall()
                    if rows:
                        msg += 'Error row {}: The email[{}] was exists in the database and cannot register.\n'.format(str(y), email)
                        error += 1
                    
            # 7
            cep         = sheet.cell(row=y, column= 7).value
            
            # 8
            address     = sheet.cell(row=y, column= 8).value

            # 9
            number      = sheet.cell(row=y, column= 9).value

            # 10
            district    = sheet.cell(row=y, column= 10).value

            # 11
            city        = sheet.cell(row=y, column= 11).value

            # 12
            state       = sheet.cell(row=y, column= 12).value

            # 13
            if error == 0:
                telefone    = sheet.cell(row=y, column = 13).value
                if email == None:
                    msg += 'Error row {}: The cell telefone[{}] wasnt preench.\n'.format(str(y), telefone)
                    error += 1
                else:
                    sql = "SELECT TELEFONE FROM CLIENT WHERE TELEFONE = '{}'".format(telefone)
                    cur.execute(sql)
                    rows = cur.fetchall()
                    if rows:
                        msg += 'Error row {}: The telefone[{}] was exists in the database and cannot register.\n'.format(str(y), telefone)
                        error += 1

            # 14
            if error == 0:
                cell    = sheet.cell(row=y, column = 14).value
                if email == None:
                    msg += 'Error row {}: The cell cell[{}] wasnt preench.\n'.format(str(y), cell)
                    error += 1
                else:
                    sql = "SELECT CELL FROM CLIENT WHERE CELL = '{}'".format(cell)
                    cur.execute(sql)
                    rows = cur.fetchall()
                    if rows:
                        msg += 'Error row {}: The cell[{}] was exists in the database and cannot register.\n'.format(str(y), cell)
                        error += 1

            if error == 0:
                sql = "INSERT INTO CLIENT(NAME, CPF, RG, BIRTH, SEX, EMAIL, CEP, ADDRESS, NUMBER, DISTRICT, CITY, STATE, TELEFONE, CELL) VALUES('{}','{}','{}','{}','{}','{}','{}','{}',{},'{}','{}','{}','{}','{}');".format(name, cpf, rg, birth, sex, email, cep, address, number, district, city, state, telefone, cell)
                #debug
                # print(sql) 
                try:
                    cur.execute(sql)
                    conf.commit()
                    msg += 'Ok row {}: The client[{}] has been successfully registered.\n'.format(str(y), name)
                except:
                    msg += 'Error row {}: The client[{}] could not be registered.\n'.format(str(y), name)
        fileName = "ImportClient_{}.txt".format(str(datetime.now()))
        file = open(fileName, "w") 
        file.write(str(msg))
        file.close
        conf.close()
        return msg
    except Exception:
        msg = 'Fail to read de file XSLX'
        return msg